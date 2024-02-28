"""
XL.py contains openpyxl code in a format usable by LabVIEW to implement Excel
functionality, without using the deprecated and buggy LabVIEW ActiveX VIs.

LabVIEW cannot use Python classes, but can access Python functions.

XL.py can open, create and save multiple workbooks at once. For each workbook,
you can read from and write to different worksheets, using various data types.

Written by Jack White.
"""

from openpyxl import Workbook, load_workbook

# workbooks to be stored in python dictionary, workbookNames are the keys
wbs = {}

##############################################################################
############################# Workbook Functions #############################
##############################################################################

# LabVIEW Function Available
def create_file(newWorkbookName):
    """
    Create a new workbook object and add it to workbooks dictionary.
    
    :param newWorkbookName: the string identifier to assign to the workbook
    :type newWorkbookName: string
    """
    
    global wbs
    wb = Workbook()
    
    try:
        # Add new wb to wbs dictionary
        wbs[newWorkbookName] = wb
    except:
        raise Exception("Could not create file") # this error hasn't occured... yet

# LabVIEW Function Available
def load_file(newWorkbookName, filePath):
    """
    Load an existing workbook and add it to workbooks dictionary.
    
    :param newWorkbookName: the name for the loading workbook
    :type newWorkbookName: string
    
    :param filePath: file path (or local name) of the workbook to be loaded
    :type filePath: string
    """
    
    global wbs
    
    try:
        wb = load_workbook(filePath)
        
        # Add new wb to wbs dictionary
        wbs[newWorkbookName] = wb
    except:
        raise Exception("File failed to load, may be open") # possible error

# LabVIEW Function Available
def list_files():
    """
    List the open workbook names.
    
    :rtype: 1D array of string types
    """
    
    return list(wbs)

# Internal Function - LabVIEW Function Not Available
def _set_active_file(workbookName):
    """
    Set the active workbook from its name in the workbooks dictionary.
    
    :param workbookName: the name of the desired workbook
    :type workbookName: string
    
    :rtype: workbook object
    """
    
    # get value (wb object) of key (workbookName) in dictionary (wbs)
    wb = wbs.get(workbookName)
    
    return wb

# LabVIEW Function Available
def save_file(workbookName, filePath):
    """
    Save the selected workbook to the path.
    
    **Note:** the wb.save() function takes the argument "Path" but if you give
    it just "file.extension", it'll save it in the same folder as XL.py.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param filePath: file path (or local name) of the workbook to be saved
    :type filePath: string
    """
    
    # set active workbook
    wb = _set_active_file(workbookName)
    
    try:
        wb.save(filePath)
    except:
        raise Exception("File failed to save, may be open") # possible error

# LabVIEW Function Available
def close_file(workbookName):
    """
    Delete the selected workbook from memory - remove it from workbooks
    dictionary.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    """
    
    global wbs
    wbs.pop(workbookName, None)

# LabVIEW Function Available
def close_all():
    """
    Delete all workbooks from memory - remove from workbooks dictionary.
    """
    
    global wbs
    wbs = {}

##############################################################################
############################ Worksheet Functions #############################
##############################################################################

# Internal Function - LabVIEW Function Not Available
def _get_active_sheet(workbookName):
    """
    Get the active worksheet's name.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :rtype: string
    """
    
    # set active workbook
    wb = _set_active_file(workbookName)
    
    worksheetName = wb.active.title
    
    return worksheetName

# Internal Function - LabVIEW Function Not Available
def _set_active_sheet(workbookName, worksheetName):
    """
    Set the active worksheet from its name.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the name of the desired worksheet
    :type worksheetName: string
    """
    
    # set active workbook
    wb = _set_active_file(workbookName)
    
    # need to return something here to be used in other functions
    wb.active = wb[worksheetName]
    
    return wb.active

# Internal Function - LabVIEW Function Not Available
def _insert_cols(workbookName, worksheetName, columnIndex, amount=1):
    """
    Insert an amount of columns before columnIndex.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param columnIndex: the column index, before which the column(s) will be
                        inserted
    :type columnIndex: int
    
    :param amount: the number of columns to insert
    :type amount: int
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # insert column before columnIndex
    ws.insert_cols(columnIndex, amount)
    
    return ws

# Internal Function - LabVIEW Function Not Available
def _insert_rows(workbookName, worksheetName, rowIndex, amount=1):
    """
    Insert an amount of rows before rowIndex.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param rowIndex: the row index, before which the row(s) will be inserted
    :type rowIndex: int
    
    :param amount: the number of columns to insert
    :type amount: int
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # insert row before rowIndex
    ws.insert_rows(rowIndex, amount)

# LabVIEW Function Available
def create_worksheet(workbookName, newWorksheetName):
    """
    Create worksheet and give it a name.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param newWorksheetName: the name of the new worksheet
    :type newWorksheetName: string
    """
    
    # set active workbook
    wb = _set_active_file(workbookName)
    
    wb.create_sheet(newWorksheetName)

# LabVIEW Function Available
def rename_worksheet(workbookName, newWorksheetName, oldWorksheetName = None):
    """
    Rename the active or selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param newWorksheetName: the new name of the worksheet
    :type newWorksheetName: string
    
    :param oldWorksheetName: the old name of the worksheet (optional). If this
    is specified, this will be the selected worksheet. Else, the default
    active worksheet will be renamed, e.g. 'Sheet'.
    :type oldWorksheetName: string
    """
    
    # set active workbook
    wb = _set_active_file(workbookName)
    
    # get current worksheet name
    wsName = oldWorksheetName or wb.active.title
    
    # get corresponding worksheet object
    ws = wb[wsName]
    
    # rename worksheet
    ws.title = newWorksheetName

# LabVIEW Function Available
def list_worksheets(workbookName):
    """
    List worksheets in selected workbook.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    """
    
    # set active workbook
    wb = _set_active_file(workbookName)
    
    return wb.sheetnames

##############################################################################
############################ Data Write Functions ############################
##############################################################################

def merge_cells_names(workbookName, worksheetName, cellName1, cellName2):
    """
    Merge cell names in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellName1: the cell in Excel format, e.g. 'A2'
    :type cellName1: string
    
    :param cellName2: the cell in Excel format, e.g. 'D2'
    :type cellName2: string
    
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # merge cells by cell name
    we.merge_cells(cellName1 + ":" + cellName2)

def unmerge_cells_names(workbookName, worksheetName, cellName1, cellName2):
    """
    Unmerge cell names in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellName1: the cell in Excel format, e.g. 'A2'
    :type cellName1: string
    
    :param cellName2: the cell in Excel format, e.g. 'D2'
    :type cellName2: string
    
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # unmerge cells by cell name
    we.unmerge_cells(cellName1 + ":" + cellName2)

def merge_cells_coords(workbookName, worksheetName, cellCoords1, cellCoords2):
    """
    Merge cell coords in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellCoords1: the cell coords in the format (col,row)
                        e.g. (1,2) = A2
    :type cellCoords1: tuple
    
    :param cellCoords2: the cell coords in the format (col,row)
                        e.g. (4,2) = D2
    :type cellCoords2: tuple
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # merge cells by cell coords
    # Note: coord[i] to match indexing in ws.cell() function,
    # e.g. in ws.cell(), (col,row) = (1,2) = A2
    we.merge_cells(start_row = cellCoords1[1],
                   start_column = cellCoords1[0],
                   end_row = cellCoords2[1],
                   end_column = cellCoords2[0])

def unmerge_cells_coords(workbookName, worksheetName, cellCoords1,
                         cellCoords2):
    """
    Unmerge cell coords in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellCoords1: the cell coords in the format (col,row)
                        e.g. (1,2) = A2
    :type cellCoords1: tuple
    
    :param cellCoords2: the cell coords in the format (col,row)
                        e.g. (4,2) = D2
    :type cellCoords2: tuple
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # unmerge cells by cell coords
    # Note: coord[i] to match indexing in ws.cell() function,
    # e.g. in ws.cell(), (col,row) = (1,2) = A2
    we.unmerge_cells(start_row = cellCoords1[1],
                     start_column = cellCoords1[0],
                     end_row = cellCoords2[1],
                     end_column = cellCoords2[0])

# LabVIEW Function Available
def write_to_cell_name(workbookName, worksheetName, cellName, value = None):
    """
    Assign value to cell name in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellName: the cell in Excel format, e.g. 'A2'
    :type cellName: string
    
    :param value: value to be inserted into cell
    :type value: float/int/str
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # assign value to cell
    ws[cellName] = value

# LabVIEW Function Available
def write_to_cell_coords(workbookName, worksheetName, cellCoords,
                         value = None):
    """
    Assign value to cell coords in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellCoords: the cell coords in the format (col,row) e.g. (1,2) = A2
    :type cellCoords: tuple
    
    :param value: value to be inserted into cell
    :type value: float/int/str
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # assign value to cell
    # Note: coord[i] to match indexing in ws.cell() function,
    # e.g. in ws.cell(), (col,row) = (1,2) = A2
    ws.cell(row=cellCoords[1], column=cellCoords[0], value=value)

# LabVIEW Function Available
def append_row(workbookName, worksheetName, row):
    """
    Append 1D array as row to selected worksheet.
    
    **Note:** this function appends row to the previous row write OR read.
    
    * For example, if all data is contained in the first 3 rows, but the 5th
    row has been read previously, then this function will append array to row
    6. This is because reading row 5 creates empty cells in memory.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param row: 1D array of data
    :type row: 1D python array of float/int/string types
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    ws.append(row)

# LabVIEW Function Available
def append_rows(workbookName, worksheetName, array):
    """
    Append 2D array as rows to selected worksheet.
    
    **Note:** this function appends rows to the previous row write OR read,
    in the same way as append_row()
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param array: 2D array of data
    :type array: 2D python array of float/int/string types
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    for row in array:
        ws.append(row)

# LabVIEW Function Available
def row_headings(workbookName, worksheetName, headings,
                 rowStart = 2, columnIndex = 1):
    """
    Insert new column before column index and fill that column with 1D array
    starting from rowStart.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param headings: 1D array of row headings
    :type headings: 1D python array of float/int/string types
    
    :param rowStart: the row to start
    :type rowStart: int
    
    :param columnIndex: the column index, before which the new column will be
                        inserted
    :type columnIndex: int
    """
    
    # set active workbook and worksheet, and insert new column
    ws = _insert_cols(workbookName, worksheetName, columnIndex)
    
    # add row headings one by one to new column, starting from rowStart
    for i, heading in enumerate(headings):
        ws.cell(row=rowStart + i, column=columnIndex, value=heading)

##############################################################################
############################ Data Read Functions #############################
##############################################################################

# LabVIEW Function Available
def read_from_cell_name(workbookName, worksheetName, cellName):
    """
    Get string value from cell name in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellName: the cell in Excel format, e.g. 'A2'
    :type cellName: string
    
    :rtype: string
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # get value from cell, str() to be LabVIEW compatible
    return str(ws[cellName].value)

# LabVIEW Function Available
def read_from_cell_coords(workbookName, worksheetName, cellCoords):
    """
    Get value from cell coords in selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param cellCoords: the cell coords in the format (col,row) e.g. (1,2) = A2
    :type cellCoords: tuple
    
    :rtype: string
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # get value from cell, str() to be LabVIEW compatible
    # Note: coord[i] to match indexing in ws.cell() function,
    # e.g. in ws.cell(), (col,row) = (1,2) = A2
    return str(ws.cell(row=cellCoords[1], column=cellCoords[0]).value)

# LabVIEW Function Available
def get_data_from_cell_names(workbookName, worksheetName, start, end):
    """
    Get 2D array of data from the start to end cell, inclusive.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param start: the start cell in Excel format, e.g. 'A2'
    :type start: string
    
    :param end: the end cell in Excel format, e.g. 'B3'
    :type end: string
    
    :rtype: 2D python array of string types
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # cell ranges can be accessed by slicing
    cellRange = ws[start:end]
    numrows = len(cellRange)
    numcols = len(cellRange[0])
    
    # initialise 2D python array, to be returned in LabVIEW, with the same
    # dimensions as cellRange
    data = [[0]*numcols for i in range(numrows)]
    
    for y, row in enumerate(cellRange):
        for x, cell in enumerate(row):
            data[y][x] = str(cell.value)
    
    return data

# LabVIEW Function Available
def get_data_from_cell_coords(workbookName, worksheetName, start, end):
    """
    Get 2D array of data from the start to end cell, inclusive.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :param start: the start cell coords in the format (col,row),
                  e.g. (1,2) = A2
    :type start: tuple
    
    :param end: the end cell coords in the format (col,row),
                e.g. (2,3) = B3
    :type end: tuple
    
    :rtype: 2D python array of string types
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # start with empty data array to be added to
    data = []
    
    # iterate over row data, with values_only = True.
    # Note: coord[i] to match indexing in iter_rows function,
    # e.g. in iter_rows: (col,row) = (1,2) = A2
    for row in ws.iter_rows(min_col = start[0],
                            min_row = start[1],
                            max_col = end[0],
                            max_row = end[1],
                            values_only = True):
        rowData = []
        for cell in row:
            rowData.append(str(cell))
        data.append(rowData)
    
    return data
 
# LabVIEW Function Available
def get_all_data(workbookName, worksheetName):
    """
    Get 2D array of all data from the selected worksheet.
    
    :param workbookName: the selected workbook name
    :type workbookName: string
    
    :param worksheetName: the selected worksheet name
    :type worksheetName: string
    
    :rtype: 2D python array of string types
    """
    
    # set active workbook and worksheet
    ws = _set_active_sheet(workbookName, worksheetName)
    
    # start with empty data array to be added to
    data = []
    
    # ws.values is a generator object which gets all data in a worksheet, but
    # needs to be converted for LabVIEW. String is the easiest type to use.
    for row in ws.values:
        rowData = []
        for cell in row:
            rowData.append(str(cell))
        data.append(rowData)
    
    return data
